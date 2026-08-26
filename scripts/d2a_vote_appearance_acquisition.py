from __future__ import annotations

import hashlib
import json
import os
import re
import time
from io import BytesIO
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

OUT = Path(os.environ.get("NEXUS_D2A_PV_OUTPUT", ".nexus-d2a-pv-public"))
SPACING = float(os.environ.get("NEXUS_REQUEST_SPACING_SECONDS", "0.65"))
TIMEOUT = float(os.environ.get("NEXUS_REQUEST_TIMEOUT_SECONDS", "30"))
USER_AGENT = "FantaDesk-Nexus-D2A-PV-Public-Acquisition/0.1"
ROLES = {"P", "D", "C", "A"}

TARGETS = [
    {
        "season": "2022/23",
        "pathSeason": "2022-23",
        "cutoffDate": "2022-08-10",
        "downloadUrl": "https://www.chiccheinformatiche.com/download/121443/",
        "sourcePage": "https://www.chiccheinformatiche.com/fantacalcio-2022-2023-il-listone-dei-giocatori-con-ruoli-e-quotazioni-di-fantagazzetta-download-excel/",
        "expectedFileName": "Quotazioni_Fantacalcio_Stagione_2022_23.xlsx",
    },
    {
        "season": "2023/24",
        "pathSeason": "2023-24",
        "cutoffDate": "2023-08-01",
        "downloadUrl": "https://www.chiccheinformatiche.com/download/123799/",
        "sourcePage": "https://www.chiccheinformatiche.com/lista-fantacalcio-quotazioni-ruoli-valori-download-excel-di-fantagazzetta-2023-2024/",
        "expectedFileName": "Quotazioni_Fantacalcio_Stagione_2023_24.xlsx",
    },
    {
        "season": "2024/25",
        "pathSeason": "2024-25",
        "cutoffDate": "2024-08-01",
        "downloadUrl": "https://www.chiccheinformatiche.com/download/131281/",
        "sourcePage": "https://www.chiccheinformatiche.com/lista-fantacalcio-quotazioni-ruoli-valori-download-excel-di-fantagazzetta-2024-2025/",
        "expectedFileName": "Quotazioni_Fantacalcio_Stagione_2024_25.xlsx",
    },
    {
        "season": "2025/26",
        "pathSeason": "2025-26",
        "cutoffDate": "2025-08-13",
        "downloadUrl": "https://www.chiccheinformatiche.com/download/133736/",
        "sourcePage": "https://www.chiccheinformatiche.com/lista-fantacalcio-quotazioni-ruoli-valori-download-excel-di-fantagazzetta-2025-2026/",
        "expectedFileName": "Quotazioni_Fantacalcio_Stagione_2025_26.xlsx",
    },
]

FINAL_SEASONS = ["2020-21", "2021-22", "2022-23", "2023-24", "2024-25", "2025-26"]

session = requests.Session()
retry = Retry(total=4, connect=4, read=4, status=4, backoff_factor=0.8, status_forcelist=(429, 500, 502, 503, 504), allowed_methods=("GET",))
session.mount("https://", HTTPAdapter(max_retries=retry))
session.headers.update({"User-Agent": USER_AGENT, "Accept-Language": "it-IT,it;q=0.9,en;q=0.7"})
last_request_at = 0.0


def sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def clean(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def request(url: str, accept: str = "*/*") -> requests.Response:
    global last_request_at
    wait = SPACING - (time.monotonic() - last_request_at)
    if wait > 0:
        time.sleep(wait)
    last_request_at = time.monotonic()
    response = session.get(url, headers={"Accept": accept}, timeout=TIMEOUT, allow_redirects=True)
    response.raise_for_status()
    return response


def save_bytes(relative: str, data: bytes) -> None:
    path = OUT / relative
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(data)


def save_json(relative: str, value) -> None:
    path = OUT / relative
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def parse_number(value):
    raw = clean(value)
    if not raw or raw in {"-", "–", "—"}:
        return None
    if "," in raw:
        raw = raw.replace(".", "").replace(",", ".")
    try:
        number = float(raw)
    except ValueError:
        return None
    if not (number == number and abs(number) != float("inf")):
        return None
    return number


def parse_listone(payload: bytes, target: dict, source_sha256: str) -> dict:
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    sheet_name = next((name for name in workbook.sheetnames if clean(name).lower() == "tutti"), None)
    if sheet_name is None:
        raise RuntimeError(f"{target['season']}: foglio Tutti assente")
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    header_index = next((idx for idx, row in enumerate(rows) if any(clean(value).lower() == "id" for value in row)), None)
    if header_index is None:
        raise RuntimeError(f"{target['season']}: header Listone assente")
    headers = [clean(value) for value in rows[header_index]]
    column = {name: idx for idx, name in enumerate(headers) if name}
    required = ["Id", "R", "Nome", "Squadra"]
    missing = [name for name in required if name not in column]
    if missing:
        raise RuntimeError(f"{target['season']}: colonne Listone mancanti {missing}")
    seen = set()
    entries = []
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        raw_id = row[column["Id"]] if column["Id"] < len(row) else None
        if raw_id is None or clean(raw_id) == "":
            continue
        try:
            fantacalcio_id = int(raw_id)
        except (TypeError, ValueError):
            raise RuntimeError(f"{target['season']}:{row_number}: Id non intero {raw_id}")
        role = clean(row[column["R"]]).upper()
        name = clean(row[column["Nome"]])
        club = clean(row[column["Squadra"]])
        if fantacalcio_id <= 0 or role not in ROLES or not name or not club:
            raise RuntimeError(f"{target['season']}:{row_number}: identità incompleta")
        if fantacalcio_id in seen:
            raise RuntimeError(f"{target['season']}: ID duplicato {fantacalcio_id}")
        seen.add(fantacalcio_id)
        entries.append({"fantacalcioId": fantacalcio_id, "name": name, "club": club, "classicRole": role, "sourceRow": row_number})
    if len(entries) < 400:
        raise RuntimeError(f"{target['season']}: Listone troppo piccolo ({len(entries)})")
    return {
        "season": target["season"],
        "cutoffAsOf": f"{target['cutoffDate']}T23:59:59.000Z",
        "sheetName": sheet_name,
        "sourcePage": target["sourcePage"],
        "expectedFileName": target["expectedFileName"],
        "sourceSha256": source_sha256,
        "entries": entries,
    }


def fantacalcio_id_from_href(href: str | None):
    match = re.search(r"/(\d+)/\d{4}-\d{2}/?(?:\?.*)?$", href or "")
    return int(match.group(1)) if match else None


def parse_final_fantacalcio(html: bytes, path_season: str) -> dict:
    soup = BeautifulSoup(html, "html.parser")
    select = soup.select_one('select[name="redaction"], select#redaction')
    selected = None
    if select:
        selected_option = select.find("option", selected=True)
        selected = selected_option.get("value") if selected_option else None
    if selected != "fantacalcio":
        raise RuntimeError(f"{path_season}: redazione selezionata {selected}")

    candidates = []
    for table in soup.find_all("table"):
        trs = table.select("tbody tr.player-row")
        if not trs:
            continue
        source_keys = [node.get("data-col-key") for node in trs[0].select("[data-col-key]") if node.get("data-col-key")]
        if "pg" not in source_keys:
            continue
        parsed = []
        malformed = []
        for tr in trs:
            cells = {node.get("data-col-key"): clean(node.get_text(" ", strip=True)) for node in tr.select("[data-col-key]") if node.get("data-col-key")}
            anchor = tr.select_one("th.player-name a")
            href = anchor.get("href") if anchor else None
            player_id = fantacalcio_id_from_href(href)
            pv = parse_number(cells.get("pg"))
            if player_id is None or pv is None or pv < 0 or pv > 38:
                malformed.append({"href": href, "pg": cells.get("pg")})
                continue
            parsed.append({
                "fantacalcioId": player_id,
                "playerName": clean(tr.select_one("th.player-name").get_text(" ", strip=True) if tr.select_one("th.player-name") else ""),
                "team": clean(cells.get("sq")),
                "appearancesWithVote": pv,
            })
        candidates.append({"sourceRowCount": len(trs), "sourceKeys": source_keys, "rows": parsed, "malformed": malformed})
    candidates.sort(key=lambda item: item["sourceRowCount"], reverse=True)
    if not candidates or len(candidates[0]["rows"]) < 100:
        raise RuntimeError(f"{path_season}: tabella Fantacalcio PV non trovata")
    best = candidates[0]
    by_id = {}
    for row in best["rows"]:
        if row["fantacalcioId"] in by_id:
            raise RuntimeError(f"{path_season}: ID finale duplicato {row['fantacalcioId']}")
        by_id[row["fantacalcioId"]] = row
    return {
        "selectedRedaction": selected,
        "sourceRowCount": best["sourceRowCount"],
        "parsedRowCount": len(best["rows"]),
        "malformedRows": len(best["malformed"]),
        "malformedExamples": best["malformed"][:20],
        "rows": best["rows"],
    }


def prior_path_season(path_season: str, lag: int) -> str:
    start = int(path_season[:4]) - lag
    return f"{start}-{str((start + 1) % 100).zfill(2)}"


def label_available_at(path_season: str) -> str:
    return f"{int(path_season[:4]) + 1}-07-15T00:00:00.000Z"


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    acquisition = {"schema": "NEXUS_D2A_PV_PUBLIC_ACQUISITION_V1", "startedAt": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()), "targets": [], "finals": []}
    listones = {}
    finals = {}

    for target in TARGETS:
        response = request(target["downloadUrl"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream,*/*")
        payload = response.content
        if not payload.startswith(b"PK\x03\x04"):
            raise RuntimeError(f"{target['season']}: risposta Listone non XLSX")
        digest = sha256(payload)
        save_bytes(f"raw/listoni/{target['expectedFileName']}", payload)
        parsed = parse_listone(payload, target, digest)
        listones[target["pathSeason"]] = parsed
        acquisition["targets"].append({"season": target["season"], "cutoffAsOf": parsed["cutoffAsOf"], "sourcePage": target["sourcePage"], "finalUrl": response.url, "sha256": digest, "rowCount": len(parsed["entries"])})

    for path_season in FINAL_SEASONS:
        url = f"https://www.fantacalcio.it/statistiche-serie-a/{path_season}/fantacalcio"
        response = request(url, "text/html,application/xhtml+xml")
        payload = response.content
        digest = sha256(payload)
        save_bytes(f"raw/final/{path_season}/page.html", payload)
        parsed = parse_final_fantacalcio(payload, path_season)
        parsed.update({"sourceUrl": url, "finalUrl": response.url, "sha256": digest})
        finals[path_season] = parsed
        acquisition["finals"].append({"season": path_season.replace("-", "/"), "sourceUrl": url, "finalUrl": response.url, "sha256": digest, "sourceRowCount": parsed["sourceRowCount"], "parsedRowCount": parsed["parsedRowCount"], "malformedRows": parsed["malformedRows"]})

    seasons = []
    for target in TARGETS:
        listone = listones[target["pathSeason"]]
        final = finals[target["pathSeason"]]
        target_by_id = {row["fantacalcioId"]: row for row in final["rows"]}
        lag1 = finals[prior_path_season(target["pathSeason"], 1)]
        lag2 = finals[prior_path_season(target["pathSeason"], 2)]
        lag1_by_id = {row["fantacalcioId"]: row for row in lag1["rows"]}
        lag2_by_id = {row["fantacalcioId"]: row for row in lag2["rows"]}
        players = []
        observed_labels = missing_labels = lag1_observed = lag2_observed = 0
        for entry in listone["entries"]:
            target_row = target_by_id.get(entry["fantacalcioId"])
            l1 = lag1_by_id.get(entry["fantacalcioId"])
            l2 = lag2_by_id.get(entry["fantacalcioId"])
            observed_labels += int(target_row is not None)
            missing_labels += int(target_row is None)
            lag1_observed += int(l1 is not None)
            lag2_observed += int(l2 is not None)
            players.append({
                "fantacalcioId": entry["fantacalcioId"],
                "playerName": entry["name"],
                "club": entry["club"],
                "classicRole": entry["classicRole"],
                "targetPv": target_row["appearancesWithVote"] if target_row else None,
                "targetMissingReason": None if target_row else "FINAL_FANTACALCIO_ROW_NOT_OBSERVED",
                "lag1Pv": l1["appearancesWithVote"] if l1 else None,
                "lag2Pv": l2["appearancesWithVote"] if l2 else None,
            })
        seasons.append({
            "targetSeason": target["season"],
            "cutoffId": f"d2a-pv:{target['pathSeason']}:preseason",
            "asOf": listone["cutoffAsOf"],
            "labelAvailableAt": label_available_at(target["pathSeason"]),
            "sourceListoneSha256": listone["sourceSha256"],
            "sourceFinalSha256": final["sha256"],
            "listonePlayers": len(players),
            "observedLabels": observed_labels,
            "missingLabels": missing_labels,
            "lag1Observed": lag1_observed,
            "lag2Observed": lag2_observed,
            "players": players,
        })

    acquisition["finishedAt"] = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
    dataset = {
        "schema": "NEXUS_D2A_PV_REPLAY_DATASET_V1",
        "createdAt": acquisition["finishedAt"],
        "targetSemantic": "Fantacalcio appearances with valid vote (PV)",
        "targetUniverse": "ARCHIVED_PRESEASON_FANTACALCIO_LISTONE",
        "identity": "FANTACALCIO_NUMERIC_ID_EXACT",
        "missingTargetPolicy": "MISSING_NOT_ZERO_EXCLUDED_FROM_MODEL_EVALUATION",
        "missingFeaturePolicy": "MISSING_RETAINED_FOR_TRAINING_PREPROCESSOR",
        "labelAvailabilityPolicy": "DERIVED_CONSERVATIVE_POST_SEASON_15_JULY",
        "marketOrQuotationFeaturesUsed": False,
        "seasons": seasons,
    }
    save_json("acquisition.json", acquisition)
    save_json("dataset.json", dataset)
    print(json.dumps({"status": "ACQUIRED", "seasons": [{k: row[k] for k in ("targetSeason", "listonePlayers", "observedLabels", "missingLabels", "lag1Observed", "lag2Observed")} for row in seasons]}, indent=2))


if __name__ == "__main__":
    main()
