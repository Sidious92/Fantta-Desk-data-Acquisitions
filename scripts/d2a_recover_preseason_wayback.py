from __future__ import annotations

import hashlib
import io
import json
import os
import time
from pathlib import Path
from urllib.parse import quote

import requests
from openpyxl import load_workbook

OUT = Path(os.environ.get("NEXUS_D2A_WAYBACK_OUTPUT", ".nexus-d2a-wayback"))
TIMEOUT = float(os.environ.get("NEXUS_REQUEST_TIMEOUT_SECONDS", "30"))
USER_AGENT = "FantaNexus-D2A-Wayback-Recovery/0.1"

TARGETS = [
    {
        "season": "2022/23",
        "download_url": "https://www.chiccheinformatiche.com/download/121443/",
        "lookup_timestamp": "20220812000000",
        "latest_allowed_timestamp": "20220813162959",
        "expected_rows": 554,
        "filename": "Quotazioni_Fantacalcio_Stagione_2022_23.xlsx",
    },
    {
        "season": "2023/24",
        "download_url": "https://www.chiccheinformatiche.com/download/123799/",
        "lookup_timestamp": "20230803000000",
        "latest_allowed_timestamp": "20230819162959",
        "expected_rows": 535,
        "filename": "Quotazioni_Fantacalcio_Stagione_2023_24.xlsx",
    },
    {
        "season": "2024/25",
        "download_url": "https://www.chiccheinformatiche.com/download/131281/",
        "lookup_timestamp": "20240803000000",
        "latest_allowed_timestamp": "20240817162959",
        "expected_rows": 480,
        "filename": "Quotazioni_Fantacalcio_Stagione_2024_25.xlsx",
    },
    {
        "season": "2025/26",
        "download_url": "https://www.chiccheinformatiche.com/download/133736/",
        "lookup_timestamp": "20250815000000",
        "latest_allowed_timestamp": "20250823162959",
        "expected_rows": 521,
        "filename": "Quotazioni_Fantacalcio_Stagione_2025_26.xlsx",
    },
]

session = requests.Session()
session.headers.update({"User-Agent": USER_AGENT})


def sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def clean(value) -> str:
    return " ".join(str(value or "").split()).strip()


def parse_count(payload: bytes) -> tuple[int, dict[str, int]]:
    wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    sheet_name = next((name for name in wb.sheetnames if clean(name).lower() == "tutti"), None)
    if sheet_name is None:
        raise RuntimeError("foglio Tutti assente")
    rows = list(wb[sheet_name].iter_rows(values_only=True))
    header_index = next((i for i, row in enumerate(rows) if any(clean(v).lower() == "id" for v in row)), None)
    if header_index is None:
        raise RuntimeError("header Id assente")
    headers = [clean(v) for v in rows[header_index]]
    col = {name: i for i, name in enumerate(headers) if name}
    for name in ("Id", "R", "Nome", "Squadra"):
        if name not in col:
            raise RuntimeError(f"colonna {name} assente")
    seen: set[int] = set()
    roles = {"P": 0, "D": 0, "C": 0, "A": 0}
    count = 0
    for row in rows[header_index + 1 :]:
        raw_id = row[col["Id"]] if col["Id"] < len(row) else None
        if raw_id is None or clean(raw_id) == "":
            continue
        player_id = int(raw_id)
        role = clean(row[col["R"]]).upper()
        name = clean(row[col["Nome"]])
        club = clean(row[col["Squadra"]])
        if player_id <= 0 or role not in roles or not name or not club:
            raise RuntimeError(f"riga identità non valida: {player_id}/{role}/{name}/{club}")
        if player_id in seen:
            raise RuntimeError(f"ID duplicato: {player_id}")
        seen.add(player_id)
        roles[role] += 1
        count += 1
    return count, roles


def availability(url: str, timestamp: str) -> dict:
    api = f"https://archive.org/wayback/available?url={quote(url, safe='')}&timestamp={timestamp}"
    r = session.get(api, timeout=TIMEOUT)
    r.raise_for_status()
    return r.json()


def raw_snapshot_url(snapshot_url: str, timestamp: str) -> str:
    marker = f"/web/{timestamp}/"
    if marker in snapshot_url:
        return snapshot_url.replace(marker, f"/web/{timestamp}id_/", 1)
    marker = f"/web/{timestamp}if_/"
    if marker in snapshot_url:
        return snapshot_url.replace(marker, f"/web/{timestamp}id_/", 1)
    return f"https://web.archive.org/web/{timestamp}id_/" + snapshot_url.split("/web/", 1)[-1].split("/", 1)[-1]


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    records = []
    failures = []
    for target in TARGETS:
        record = {k: target[k] for k in ("season", "download_url", "lookup_timestamp", "latest_allowed_timestamp", "expected_rows", "filename")}
        try:
            payload = availability(target["download_url"], target["lookup_timestamp"])
            closest = payload.get("archived_snapshots", {}).get("closest") or {}
            if closest.get("available") is not True or str(closest.get("status")) != "200":
                raise RuntimeError("nessuna cattura Wayback 200 disponibile")
            capture_ts = str(closest.get("timestamp") or "")
            if len(capture_ts) != 14 or not capture_ts.isdigit():
                raise RuntimeError(f"timestamp cattura non valido: {capture_ts}")
            if capture_ts > target["latest_allowed_timestamp"]:
                raise RuntimeError(f"cattura post-kickoff non ammessa: {capture_ts}")
            snap = str(closest.get("url") or "")
            raw_url = raw_snapshot_url(snap, capture_ts)
            r = session.get(raw_url, timeout=TIMEOUT, allow_redirects=True)
            r.raise_for_status()
            data = r.content
            if not data.startswith(b"PK\x03\x04"):
                raise RuntimeError(f"payload non XLSX: content-type={r.headers.get('content-type')}")
            row_count, roles = parse_count(data)
            if row_count != target["expected_rows"]:
                raise RuntimeError(f"fingerprint row-count mismatch: {row_count} != {target['expected_rows']}")
            path = OUT / "listoni" / target["filename"]
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes(data)
            record.update({
                "status": "PASS",
                "capture_timestamp": capture_ts,
                "snapshot_url": snap,
                "raw_snapshot_url": raw_url,
                "final_url": r.url,
                "byte_length": len(data),
                "sha256": sha256(data),
                "row_count": row_count,
                "role_counts": roles,
            })
        except Exception as exc:
            record.update({"status": "FAIL", "error": f"{type(exc).__name__}: {exc}"})
            failures.append({"season": target["season"], "error": record["error"]})
        records.append(record)
        time.sleep(1.0)
    manifest = {
        "schema": "NEXUS_D2A_PRESEASON_WAYBACK_RECOVERY_V1",
        "generated_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "policy": {
            "pre_kickoff_capture_required": True,
            "xlsx_magic_required": True,
            "expected_historical_row_counts": {t["season"]: t["expected_rows"] for t in TARGETS},
            "missing_or_mismatch": "FAIL_CLOSED",
            "tls_verification_disabled": False,
        },
        "records": records,
        "status": "PASS" if not failures else "FAIL_CLOSED",
        "failures": failures,
    }
    (OUT / "manifest.json").write_text(json.dumps(manifest, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps(manifest, indent=2, ensure_ascii=False))
    if failures:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
